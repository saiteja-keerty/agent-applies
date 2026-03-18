"""
╔══════════════════════════════════════════════════════╗
║           EMAIL SENDER FOR JOB APPLICATIONS          ║
║    Reads Excel file and sends personalized emails    ║
║         to hiring teams with full context            ║
╚══════════════════════════════════════════════════════╝

HOW TO USE:
  1. Run job_hunter.py (creates job_applications.xlsx)
  2. Open job_applications.xlsx and mark jobs as "Applied" 
     in the Status column (the ones you want to send emails to)
  3. python send_emails.py
  4. Preview emails BEFORE sending (DRY_RUN = True)
  5. Change DRY_RUN = False when ready to actually send

GMAIL SETUP (recommended):
  1. Enable 2-Step Verification: https://myaccount.google.com/security
  2. Get App Password: https://myaccount.google.com/apppasswords
  3. Paste the app password (16 chars) into SENDER_PASSWORD below
  4. Set SENDER_EMAIL = "yourname@gmail.com"

ALTERNATIVE: Use your own email provider (Outlook, Yahoo, etc.)
"""

import smtplib
import openpyxl
import time
import os
import base64
import json
import re
import requests
from datetime import datetime
from bs4 import BeautifulSoup
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, use system env vars

# ─────────────────────────────────────────────
#  YOUR EMAIL CONFIGURATION — EDIT THESE
# ─────────────────────────────────────────────

SENDER_EMAIL    = "saiteja.keerty@gmail.com"      # Your Gmail or email
SENDER_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "YOUR_APP_PASSWORD_HERE")       # Gmail App Password (16 chars)
SENDER_NAME     = "Saiteja Keerty"                   # Your name
SENDER_PHONE    = "203-901-6018"              # Your phone number
RESUME_FILE     = "Saiteja_keerty_Resume.pdf" # Your resume file (paste your resume as PDF in this folder)

# Email settings
SMTP_HOST       = "smtp.gmail.com"             # Gmail SMTP
SMTP_PORT       = 587                          # TLS port
DRY_RUN         = False                        # Set to False to ACTUALLY send emails
MAX_EMAILS      = 5                            # Limit number of emails to send (set to None for all)
                                                # First run with True to preview!

# SEND ALL JOBS: Automatically sends to all jobs in Excel

# Input/Output files
PROJECT_FOLDER  = os.path.dirname(os.path.abspath(__file__))  # Script's folder
INPUT_FILE      = os.path.join(PROJECT_FOLDER, "job_applications.xlsx")  # From job_hunter.py
PROFESSIONALS_FILE = os.path.join(PROJECT_FOLDER, "professional_contacts.json")  # Professional contacts
OUTPUT_FILE     = os.path.join(PROJECT_FOLDER, "send_log.txt")  # Log of what was sent

# DUPLICATE PREVENTION
SENT_LOG_FILE   = os.path.join(PROJECT_FOLDER, "sent_emails.json")  # Track sent emails

# API Keys for email finding
HUNTER_API_KEY  = os.getenv("HUNTER_API_KEY", "YOUR_HUNTER_API_KEY_HERE")  # Hunter.io API key

# Email template — this will be used for ALL emails
# {company}, {job_title}, {job_url}, {cover_letter} will be replaced with actual data
EMAIL_BODY_TEMPLATE = """Hello,

I hope you are doing well.

I am writing to express my strong interest in the {job_title} position at {company}. With dedicated experience in data engineering, I have worked for both startups and multinational companies (MNCs) where I specialized in building systems from scratch and developing end-to-end pipelines. My background includes successfully delivering R&D projects and proofs of concept (POCs).

I bring significant expertise in:
• Databricks & Cloud platforms (AWS, Azure, GCP)
• ETL pipelines & data architecture
• Languages: Python, SQL, Data Science and AI tools
• Infrastructure tools: Airflow, Terraform
• Startup & MNC experience in R&D projects and building POCs

I am confident that my skills and experience align closely with what you're looking for. Please find my resume attached for your review.

I am available to start immediately and would welcome the opportunity to discuss how my background can contribute to your team.

Thank you for your time and consideration. I look forward to hearing from you.

Best regards,
{sender_name}
{sender_phone}
{sender_email}

---
Sent using own Job Agent - open source built by me - sorry if spammed you!:
"""

# Personalized email template for professionals
PERSONAL_EMAIL_TEMPLATE = """Hi {person_name},

I hope you are doing well.

I'm reaching out because I admire the work you've been doing at {company}. I'm a data engineer with 5+ years of experience in building scalable data pipelines and ETL workflows, and I'm particularly interested in opportunities to contribute to your team.

My background includes:
• Designing scalable Terraform modules for Databricks, AWS, and Redshift
• Building end-to-end ETL pipelines using Airflow, Spark, and modern cloud platforms
• Implementing data observability solutions and governance policies
• Successfully delivering R&D projects and POCs for startups and multinational companies
• Experience with AWS, GCP, Azure, and multiple data platforms

I'd love to discuss how my experience can benefit {company}. Please find my resume attached.

I'm available for a quick call or meeting at your convenience.

Best regards,
{sender_name}
{sender_phone}
{sender_email}

---
Sent using own Job Agent - open source built by me"""

# ─────────────────────────────────────────────
#  📧  EMAIL SENDER LOGIC
# ─────────────────────────────────────────────

def load_professional_contacts() -> dict:
    """Load professional contacts from JSON file."""
    if os.path.exists(PROFESSIONALS_FILE):
        try:
            with open(PROFESSIONALS_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}


def load_sent_emails():
    """Load previously sent emails to prevent duplicates."""
    if os.path.exists(SENT_LOG_FILE):
        try:
            with open(SENT_LOG_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_sent_email(company, job_title, email):
    """Save sent email to prevent future duplicates."""
    sent_emails = load_sent_emails()
    key = f"{company.lower()}|{job_title.lower()}|{email.lower()}"
    sent_emails[key] = {
        "company": company,
        "job_title": job_title,
        "email": email,
        "sent_date": datetime.now().isoformat()
    }
    
    with open(SENT_LOG_FILE, 'w') as f:
        json.dump(sent_emails, f, indent=2)


def is_email_already_sent(company, job_title, email):
    """Check if email was already sent to prevent duplicates."""
    sent_emails = load_sent_emails()
    key = f"{company.lower()}|{job_title.lower()}|{email.lower()}"
    return key in sent_emails


def update_excel_status(row_number, status):
    """Update the status column in Excel for sent emails."""
    try:
        wb = openpyxl.load_workbook(INPUT_FILE)
        ws = wb["Applications"]
        ws.cell(row=row_number, column=1, value=status)
        wb.save(INPUT_FILE)
    except Exception as e:
        print(f"    ⚠️  Could not update Excel status: {e}")


def read_excel(filename: str) -> list[dict]:
    """Read job data from Excel file."""
    wb = openpyxl.load_workbook(filename)
    ws = wb["Applications"]
    
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=False):  # Skip header
        job = {
            "status":      row[0].value,
            "company":     row[1].value,
            "job_title":   row[2].value,
            "location":    row[3].value,
            "source":      row[4].value,
            "url":         row[5].value,
            "email_1":     row[6].value,
            "email_2":     row[7].value,
            "email_3":     row[8].value,
            "email_4":     row[9].value,
            "email_5":     row[10].value,
            "cover_letter":row[11].value,
            "date_found":  row[12].value,
            "notes":       row[13].value,
            "excel_row":   row[0].row,  # Track which row for later logging
        }
        jobs.append(job)
    
    return jobs


def build_email_body(job: dict, sender_name: str, sender_email: str, sender_phone: str) -> str:
    """Build full email body with context."""
    body = EMAIL_BODY_TEMPLATE.format(
        job_title=job.get("job_title", ""),
        company=job.get("company", ""),
        sender_name=sender_name,
        sender_email=sender_email,
        sender_phone=sender_phone,
    )
    return body


def build_personal_email_body(person_name: str, company: str, sender_name: str, sender_email: str, sender_phone: str) -> str:
    """Build personalized email body for professionals."""
    body = PERSONAL_EMAIL_TEMPLATE.format(
        person_name=person_name,
        company=company,
        sender_name=sender_name,
        sender_email=sender_email,
        sender_phone=sender_phone,
    )
    return body


# Additional email finding functions (copied from job_hunter.py)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

def get_domain_from_url(url: str) -> str:
    """Extract domain from a URL."""
    if not url:
        return ""
    url = url.replace("https://", "").replace("http://", "").replace("www.", "")
    return url.split("/")[0].strip()


def guess_emails(company: str, company_url: str) -> list[str]:
    """
    Generate a list of likely email addresses for a company.
    These are guesses — you verify manually before sending.
    """
    domain = get_domain_from_url(company_url)
    if not domain or len(domain) < 4:
        # Try to build domain from company name
        clean = re.sub(r'[^a-zA-Z0-9]', '', company.lower().split()[0])
        domain = f"{clean}.com"

    # Generate HR/careers emails (these are safe generic guesses)
    emails = [
        f"careers@{domain}",
        f"hr@{domain}",
      #  f"recruiting@{domain}",
      #  f"hiring@{domain}",
      #  f"jobs@{domain}",
    ]
    return emails


def scrape_emails_from_url(url: str) -> list[str]:
    """
    Scrape emails from a given URL.
    """
    if not url:
        return []
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text()
        # Find all emails using regex
        email_pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
        emails = re.findall(email_pattern, text)
        # Filter valid emails
        valid_emails = [email for email in emails if verify_email_syntax(email)]
        # Remove duplicates and return
        return list(set(valid_emails))
    except Exception as e:
        print(f"    ⚠️  Error scraping {url}: {e}")
        return []


def get_emails_from_hunter(domain: str) -> list[str]:
    """
    Get emails from Hunter.io domain search.
    """
    if not HUNTER_API_KEY or not domain:
        return []
    try:
        url = f"https://api.hunter.io/v2/domain-search?domain={domain}&api_key={HUNTER_API_KEY}"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        emails = []
        for email_data in data.get("data", {}).get("emails", []):
            email = email_data.get("value")
            if email and verify_email_syntax(email):
                emails.append(email)
        return emails[:10]  # limit to 10
    except Exception as e:
        print(f"    ⚠️  Hunter.io error for {domain}: {e}")
        return []


def search_linkedin_professionals(job_title: str, company: str) -> list[str]:
    """
    Return empty list - LinkedIn searches are too slow/unreliable.
    Professional contacts will come from job_hunter.py scraping instead.
    """
    return []


def search_google_for_emails(job_title: str, company: str) -> list[str]:
    """
    Return empty list - Google searches are too slow/unreliable for sending.
    Professional contacts will come from job_hunter.py scraping instead.
    """
    return []


def search_professional_networks(company: str, job_title: str) -> list[str]:
    """
    Return empty list - Website scraping is too slow/unreliable.
    Professional contacts will come from job_hunter.py scraping and guesses.
    """
    return [][:10]


def verify_email_syntax(email: str) -> bool:
    """Basic syntax check."""
    pattern = r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def find_additional_emails(job: dict) -> list[str]:
    """Find additional emails for a job - fast version without slow searches."""
    company = job.get("company", "")
    company_url = job.get("company_url", job.get("url", ""))

    # For now, just use guessed emails (careers@, hr@, etc)
    # Professional contacts will come from job_hunter.py professional_contacts.json
    guessed = guess_emails(company, company_url)
    
    return guessed


def send_email(recipient_email: str, job: dict, sender_email: str, sender_password: str, 
               sender_name: str, sender_phone: str, dry_run: bool = True, person_name: str = None, is_personal: bool = False) -> dict:
    """Send an individual email."""
    
    company = job["company"]
    if is_personal and person_name:
        subject = f"Connecting: Data Engineer at {company}"
        body = build_personal_email_body(person_name, company, sender_name, sender_email, sender_phone)
    else:
        subject = f"Application: {job['job_title']} at {job['company']} - 5 years"
        body = build_email_body(job, sender_name, sender_email, sender_phone)
    
    result = {
        "company": job["company"],
        "job_title": job["job_title"],
        "recipient": recipient_email,
        "person_name": person_name,
        "is_personal": is_personal,
        "status": "PREVIEW" if dry_run else "SENT",
        "timestamp": datetime.now().isoformat(),
        "error": None,
    }
    
    if dry_run:
        print(f"\n{'='*70}")
        print(f"DRY RUN — Would send to: {recipient_email}" + (f" ({person_name})" if person_name else ""))
        print(f"{'='*70}")
        print(f"TO: {recipient_email}")
        print(f"SUBJECT: {subject}")
        print(f"FROM: {sender_name} <{sender_email}>")
        if os.path.exists(RESUME_FILE):
            print(f"ATTACHMENT: {RESUME_FILE}")
        print(f"\n{body}")
        print(f"{'='*70}\n")
        return result
    
    # Actually send email
    try:
        msg = MIMEMultipart()
        msg["From"] = f"{sender_name} <{sender_email}>"
        msg["To"] = recipient_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))
        
        # Attach resume if it exists
        if os.path.exists(RESUME_FILE):
            with open(RESUME_FILE, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename= {RESUME_FILE}")
            msg.attach(part)
        
        person_info = f" ({person_name})" if person_name else ""
        print(f"  Sending to {recipient_email}{person_info}...", end=" ")
        
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        
        print("Sent!")
        return result
    
    except Exception as e:
        print(f"Failed: {e}")
        result["status"] = "FAILED"
        result["error"] = str(e)
        return result


def log_results(results: list[dict], filename: str):
    """Save email sending log."""
    with open(filename, "w") as f:
        f.write("EMAIL SENDING LOG\n")
        f.write("="*70 + "\n")
        f.write(f"Generated: {datetime.now()}\n\n")
        
        for result in results:
            f.write(f"Company: {result['company']}\n")
            f.write(f"Job: {result['job_title']}\n")
            f.write(f"Recipient: {result['recipient']}\n")
            if result.get('person_name'):
                f.write(f"Person: {result['person_name']}\n")
            if result.get('is_personal'):
                f.write(f"Type: Personalized Email\n")
            f.write(f"Status: {result['status']}\n")
            if result.get('error'):
                f.write(f"Error: {result['error']}\n")
            f.write(f"Time: {result['timestamp']}\n")
            f.write("-"*70 + "\n\n")
    
    print(f"  Log saved to: {filename}")


# ─────────────────────────────────────────────
#  🚀  MAIN RUNNER
# ─────────────────────────────────────────────

def main():
    print("\n" + "="*70)
    print("  EMAIL SENDER FOR JOB APPLICATIONS")
    print("="*70 + "\n")
    
    # Safety check
    if DRY_RUN:
        print("  WARNING: DRY RUN MODE — Just showing what WOULD be sent")
        print("  Set DRY_RUN = False in the script to actually send emails\n")
    else:
        confirm = input("  WARNING: REAL SEND MODE — Actually send emails? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("  Cancelled.")
            return
    
    # Load jobs
    print(f"  Reading {INPUT_FILE}...")
    jobs = read_excel(INPUT_FILE)
    
    if not jobs:
        print(f"  WARNING: No jobs found in Excel!")
        print(f"  Please run job_hunter.py to scrape jobs first.")
        return
    
    print(f"  Found {len(jobs)} jobs in Excel")
    if MAX_EMAILS:
        jobs = jobs[:MAX_EMAILS]
        print(f"  Limiting to first {MAX_EMAILS} jobs")
    print(f"  Checking for duplicate emails...\n")
    
    all_results = []
    
    # Load professional contacts
    print(f"  Loading professional contacts...")
    professional_contacts = load_professional_contacts()
    print(f"  Found professional contacts for {len(professional_contacts)} companies\n")
    
    # Send emails to ALL jobs
    for i, job in enumerate(jobs, 1):
        print(f"  [{i}/{len(jobs)}] {job['company']} — {job['job_title']}")
        
        # Determine target email(s) to try
        emails_to_try = [job["email_1"], job["email_2"], job["email_3"], job["email_4"], job["email_5"]]
        emails_to_try = [email for email in emails_to_try if email]  # Remove empty emails
        
        # Find additional emails from Google, LinkedIn, etc.
        additional_emails = find_additional_emails(job)
        emails_to_try.extend(additional_emails)
        
        # Remove duplicates
        emails_to_try = list(set(emails_to_try))
        
        if not emails_to_try:
            print(f"    WARNING: No email found, skipping...")
        else:
            # Try each email address until one succeeds
            email_sent = False
            for target_email in emails_to_try:
                # Check for duplicates
                if is_email_already_sent(job["company"], job["job_title"], target_email):
                    print(f"    WARNING: Already sent to {target_email}, trying next...")
                    continue
                
                print(f"    Trying: {target_email}")
                result = send_email(
                    recipient_email=target_email,
                    job=job,
                    sender_email=SENDER_EMAIL,
                    sender_password=SENDER_PASSWORD,
                    sender_name=SENDER_NAME,
                    sender_phone=SENDER_PHONE,
                    dry_run=DRY_RUN,
                    is_personal=False
                )
                
                # Mark as sent if successful
                if result["status"] == "SENT" and not DRY_RUN:
                    save_sent_email(job["company"], job["job_title"], target_email)
                    update_excel_status(job["excel_row"], "Sent")
                    email_sent = True
                    print(f"    SUCCESS: Email sent to {target_email}")
                    break
                else:
                    print(f"    FAILED: Could not send to {target_email}")
            
            if not email_sent and not DRY_RUN:
                print(f"    ERROR: Failed to send to any email address for this job")
                result = {
                    "company": job["company"],
                    "job_title": job["job_title"],
                    "recipient": " | ".join(emails_to_try),
                    "person_name": None,
                    "is_personal": False,
                    "status": "FAILED",
                    "timestamp": datetime.now().isoformat(),
                    "error": "All email addresses failed",
                }
                all_results.append(result)
        
        # Send personalized emails to professional contacts from this job
        job_professionals = job.get("professional_contacts", [])
        if job_professionals:
            print(f"    📝 Sending personalized emails to {len(job_professionals)} professionals...")
            for prof in job_professionals:
                prof_email = prof.get("email") if isinstance(prof, dict) else prof
                prof_name = prof.get("name") if isinstance(prof, dict) else "Professional"
                
                if not prof_email:
                    continue
                
                # Check for duplicates
                if is_email_already_sent(job["company"], job["job_title"], prof_email):
                    print(f"      Already sent to {prof_email}, skipping...")
                    continue
                
                print(f"      Sending to {prof_name} ({prof_email})...")
                result = send_email(
                    recipient_email=prof_email,
                    job=job,
                    sender_email=SENDER_EMAIL,
                    sender_password=SENDER_PASSWORD,
                    sender_name=SENDER_NAME,
                    sender_phone=SENDER_PHONE,
                    dry_run=DRY_RUN,
                    person_name=prof_name,
                    is_personal=True
                )
                
                all_results.append(result)
                
                if result["status"] == "SENT" and not DRY_RUN:
                    save_sent_email(job["company"], job["job_title"], prof_email)
                    print(f"      SUCCESS: Personalized email sent to {prof_name}")
                    time.sleep(1)  # Rate limit
                else:
                    print(f"      FAILED: Could not send to {prof_name}")
        
        # Send personalized emails to professional contacts
        company_lower = job.get('company', '').lower()
        prof_contacts = professional_contacts.get(company_lower, {}).get('contacts', [])
        
        if prof_contacts:
            print(f"    📝 Sending personalized emails to {len(prof_contacts)} professionals...")
            for prof in prof_contacts:
                prof_email = prof.get("email") if isinstance(prof, dict) else prof
                prof_name = prof.get("name") if isinstance(prof, dict) else "Professional"
                
                if not prof_email or not isinstance(prof_email, str):
                    continue
                
                # Check for duplicates
                if is_email_already_sent(job["company"], job["job_title"], prof_email):
                    print(f"      Already sent to {prof_email}, skipping...")
                    continue
                
                print(f"      Sending to {prof_name} ({prof_email})...")
                result = send_email(
                    recipient_email=prof_email,
                    job=job,
                    sender_email=SENDER_EMAIL,
                    sender_password=SENDER_PASSWORD,
                    sender_name=SENDER_NAME,
                    sender_phone=SENDER_PHONE,
                    dry_run=DRY_RUN,
                    person_name=prof_name,
                    is_personal=True
                )
                
                all_results.append(result)
                
                if result["status"] == "SENT" and not DRY_RUN:
                    save_sent_email(job["company"], job["job_title"], prof_email)
                    print(f"      SUCCESS: Personalized email sent to {prof_name}")
                    time.sleep(1)  # Rate limit
                else:
                    print(f"      FAILED: Could not send to {prof_name}")
        
        if not DRY_RUN:
            time.sleep(2)  # Rate limit — don't spam
    
    # Send emails to professional contacts not yet assigned to a job
    print(f"\n  📧 Sending to additional professional contacts...")
    for company_key, prof_data in professional_contacts.items():
        for prof in prof_data.get('contacts', []):
            prof_email = prof.get("email") if isinstance(prof, dict) else prof
            prof_name = prof.get("name") if isinstance(prof, dict) else "Professional"
            company_name = prof_data.get('company', company_key)
            
            if not prof_email or not isinstance(prof_email, str):
                continue
            
            # Check for duplicates across all jobs
            already_sent = False
            for job in jobs:
                if is_email_already_sent(job["company"], job["job_title"], prof_email):
                    already_sent = True
                    break
            
            if already_sent:
                continue
            
            # Create a generic job info for this contact
            generic_job = {
                "company": company_name,
                "job_title": "Data Engineering Role",
            }
            
            print(f"    Sending to {prof_name} at {company_name} ({prof_email})...")
            result = send_email(
                recipient_email=prof_email,
                job=generic_job,
                sender_email=SENDER_EMAIL,
                sender_password=SENDER_PASSWORD,
                sender_name=SENDER_NAME,
                sender_phone=SENDER_PHONE,
                dry_run=DRY_RUN,
                person_name=prof_name,
                is_personal=True
            )
            
            all_results.append(result)
            
            if result["status"] == "SENT" and not DRY_RUN:
                save_sent_email(company_name, "Data Engineering Role", prof_email)
                print(f"    SUCCESS: Personalized email sent to {prof_name}")
                time.sleep(1)
            
            if not DRY_RUN:
                time.sleep(1)
    
    # Summary
    print("\n" + "="*70)
    if DRY_RUN:
        print(f"  PREVIEW COMPLETE — {len(all_results)} emails shown above")
        print("\n  NEXT STEPS:")
        print("  1. Review the email previews above")
        print("  2. If they look good, change DRY_RUN = False")
        print("  3. Run this script again to send for real")
    else:
        sent = len([r for r in all_results if r["status"] == "SENT"])
        failed = len([r for r in all_results if r["status"] == "FAILED"])
        print(f"  Sent to companies: {sent} emails")
        if failed:
            print(f"  Failed: {failed} emails")
    
    print("="*70 + "\n")
    
    # Log results
    log_results(all_results, OUTPUT_FILE)


if __name__ == "__main__":
    main()
