"""
╔══════════════════════════════════════════════════════╗
║        FREE JOB HUNTER + EMAIL FINDER AGENT          ║
║  Scrapes jobs, finds emails, writes to Excel         ║
║  Cost: $0  |  RAM: works on 8GB                      ║
╚══════════════════════════════════════════════════════╝

SETUP (run once):
  pip install playwright beautifulsoup4 requests openpyxl groq google-search-results
  playwright install chromium

FREE API KEYS NEEDED:
  - Groq (AI): https://console.groq.com  (free, no credit card)
  - Optional: Hunter.io free tier (150/mo): https://hunter.io
  - Optional: Google Search API: https://serpapi.com or https://www.programmablesearchengine.com
  - LinkedIn Profile: https://www.linkedin.com (for scraping professionals)

HOW TO RUN:
  python job_hunter.py

OUTPUT:
  job_applications.xlsx  — ready to review and send manually
"""

import time
import re
import json
import random
import datetime
import requests
import os
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
#  ⚙️  YOUR CONFIGURATION — EDIT THESE
# ─────────────────────────────────────────────

YOUR_NAME       = "Naga Keerty"
YOUR_EMAIL      = "saiteja.keerty@gmail.com"
YOUR_PHONE      = "+1 203-901-6018"
YOUR_LINKEDIN   = "https://www.linkedin.com/in/saiteja-keerty669073190/"

# Job search settings
JOB_TITLES      = ["data engineer", "Devops engineer", "platform data engineer", "aws data engineer"]  # add more titles as needed
LOCATION        = "Remote" or "New York" or "USA"  # or "New York" or "London"
MAX_JOBS        = 50      # how many jobs to collect per run

# Your resume as plain text (paste it here)
RESUME_TEXT = """
Sai Teja Keerty 
Houston | +1-203-901-6018  | saiteja.keerty@gmail.com | Linkedin | Git 
Data engineer with over 5 years of experience designing, building, and maintaining scalable data pipelines and ETL workflows for diverse data 
environments. Skilled in AWS, GCP & Azure services, Terraform, Spark, and Python to ensure data accuracy, consistency, and reproducibility. 
Proven track record optimizing data processes, cutting transformation times, and improving system performance and reliability.  
WORK EXPERIENCE 
GOODRX | Data Engineer Apr 2025 - Feb 2026 
• Designed scalable, reusable Terraform modules to provision and manage Databricks, AWS, and Redshift infrastructure, standardizing 
environments and reducing deployment/setup time across dev, QA, and prod. 
• Integrated Databricks with Monte Carlo to implement real-time data observability and alerting, resolving token scope and permission 
issues and enabling proactive detection of data quality incidents. 
• Defined and enforced Databricks cluster governance policies via Terraform (runtime versions, Python package controls, access standards), 
eliminating manual configuration drift and ensuring consistent compliance at scale. 
• Built and optimized end-to-end ETL pipelines using Airflow, Databricks, and Redshift, including incremental load strategies and 
performance tuning, improving reliability and scalability for large healthcare and marketing datasets. 
• Delivered CMI Provider Level Data (PLD) exports aligned to contractual and reporting standards, ensuring consistent, accurate, and 
timely partner data delivery with no downstream rework. 
• Collaborated in R&D initiatives to design a POC on RAG-based chatbot for answering policy and internal documentation queries, 
integrating secure data retrieval pipelines and enterprise compliance controls. 
TOWN FAIR TIRE | Data Engineer Jan 2024 - Dec 2024 
• Designed and optimized Spark data pipelines and migrated key processes from SQL to PySpark, achieving a 50% reduction in web app display 
time.  
• Orchestrated multiple concurrent processes on Spark, effectively mitigating SQL resource deadlocks and improving system performance. 
• Utilized analytical thinking to analyze customer data using SQL, leading to the development of a rewards program based on annual visits and a 
3% increase in customer retention.  
• Researched and engineered the deployment of Apache Kafka and Spark, successfully setting up development, test, and production environments 
from scratch.  
• Implemented Git, Grafana, Prometheus, and Apache Airflow independently, enabling comprehensive monitoring, alerting, task scheduling, and 
automated backups, streamlining workflows and enhancing system reliability.  
Tata Consultancy Services | System Engineer (Data Engineer) Jul 2021 - Jul 2022 
• Architected and deployed data ingestion pipelines using AWS Lambda, AWS Data Pipeline, and Kinesis Streams to process high-volume 
streaming data.  
• Developed ETL workflows for structured and semi-structured data, integrating Amazon RDS, S3, and Redshift, improving data accessibility by  
40%.  
• Built Python-based automation scripts to process and clean large datasets, reducing data transformation time by 50%. 
• Led the development of data governance and security measures using AWS Key Management Service (KMS) and AWS Glue Data Catalog. 
• Collaborated with DevOps teams to optimize CI/CD pipelines for data engineering workflows using Terraform and AWS CloudFormation. 
Tata Consultancy Services | Assistant System Engineer Jul 2019 - Jul 2021 
• Designed and implemented scalable data models in Amazon Redshift, improving query performance and reporting speed by 20%. 
• Developed and maintained complex SQL queries, stored procedures, and triggers to support data analytics and reporting needs. 
• Automated ETL workflows using Python and Apache Airflow, reducing manual intervention and enhancing reliability. 
• Ensured data integrity and quality by implementing validation rules and monitoring pipelines through AWS CloudWatch and Prometheus. 
CERTIFICATIONS 
•Certified GCP Data engineer (PDE, GCP): Nov 2022 
EDUCATION 
University of New Haven 
MS, Data Science (GPA: 3.7/4) 
Coursework: Machine Learning for data science, Data 
Engineering, Cloud Data Pipelines  
Pragati Engineering College  
Bachelor of technology, Mechanical Engineering 
SKILLS  
Aug 2022 - May 2024 
Connecticut, CT 
Jul 2015 - May 2019 
Andhra Pradesh, India 
• Programming & Scripting: Python, SQL, Shell Scripting, PySpark, Langchain, Langraph, RAG, PineCone, AI Agent 
• Cloud & Data Services: AWS Data Pipeline, AWS Lambda, AWS Step Functions, Amazon S3, fivetran, Amazon Kinesis, Amazon RDS 
(MySQL, PostgreSQL)  
• ETL, Data Processing & Pipeline Design: Apache Spark, Airflow, Pandas, NumPy, Data Integration Techniques, Data Pipeline Design 
• Data Warehousing & Databases: Amazon Redshift, MySQL, PostgreSQL, DynamoDB 
• DevOps, Monitoring & Security: Terraform, CI/CD (GitHub, Jenkins), Docker, Kubernetes, codefresh, codecov, AWS CloudWatch, 
Prometheus, Grafana, monte carlo  
• Methodologies & Tools: Analytical Thinking, Agile Methodologies, Microsoft Office Suite
"""

# AI Settings (Groq is free)
GROQ_API_KEY    = os.getenv("GROQ_API_KEY", "YOUR_GROQ_API_KEY_HERE")   # get at console.groq.com
USE_AI          = False   # set False to skip cover letter generation

# Hunter.io for finding professional emails (free tier: 150/month)
HUNTER_API_KEY  = "b54fa43ede0d69f76a2e2bdc2f565e092252e152"  # get at hunter.io, optional

# Google Search API (optional - for finding emails via Google)
GOOGLE_API_KEY  = ""  # get from serpapi.com or programmablesearchengine.com
SERP_API_KEY    = ""  # SerpAPI key for Google searches

# Output file
OUTPUT_FILE     = "job_applications.xlsx"


# ─────────────────────────────────────────────
#  🔍  JOB SCRAPERS
# ─────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

def scrape_remotive(job_title: str, max_results: int = 10) -> list[dict]:
    """Scrape Remotive.com — great for remote tech jobs, no auth needed."""
    print(f"  🌐 Scraping Remotive for: {job_title}")
    jobs = []
    try:
        query = job_title.replace(" ", "+")
        url = f"https://remotive.com/api/remote-jobs?search={query}&limit={max_results}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        data = resp.json()
        for job in data.get("jobs", [])[:max_results]:
            jobs.append({
                "title":       job.get("title", ""),
                "company":     job.get("company_name", ""),
                "location":    job.get("candidate_required_location", "Remote"),
                "url":         job.get("url", ""),
                "description": BeautifulSoup(job.get("description", ""), "html.parser").get_text()[:500],
                "source":      "Remotive",
                "date_found":  datetime.date.today().isoformat(),
                "company_url": job.get("company_url", ""),
            })
    except Exception as e:
        print(f"    ⚠️  Remotive error: {e}")
    return jobs


def scrape_arbeitnow(job_title: str, max_results: int = 10) -> list[dict]:
    """Scrape Arbeitnow — free API, no key needed, EU + remote jobs."""
    print(f"  🌐 Scraping Arbeitnow for: {job_title}")
    jobs = []
    try:
        query = job_title.replace(" ", "+")
        url = f"https://www.arbeitnow.com/api/job-board-api?search={query}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        data = resp.json()
        for job in data.get("data", [])[:max_results]:
            jobs.append({
                "title":       job.get("title", ""),
                "company":     job.get("company_name", ""),
                "location":    job.get("location", "Remote"),
                "url":         job.get("url", ""),
                "description": job.get("description", "")[:500],
                "source":      "Arbeitnow",
                "date_found":  datetime.date.today().isoformat(),
                "company_url": "",
            })
    except Exception as e:
        print(f"    ⚠️  Arbeitnow error: {e}")
    return jobs


def scrape_jobicy(job_title: str, max_results: int = 10) -> list[dict]:
    """Scrape Jobicy — free remote job board API."""
    print(f"  🌐 Scraping Jobicy for: {job_title}")
    jobs = []
    try:
        query = job_title.replace(" ", "%20")
        url = f"https://jobicy.com/api/v2/remote-jobs?count={max_results}&keyword={query}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        data = resp.json()
        for job in data.get("jobs", [])[:max_results]:
            jobs.append({
                "title":       job.get("jobTitle", ""),
                "company":     job.get("companyName", ""),
                "location":    job.get("jobGeo", "Remote"),
                "url":         job.get("url", ""),
                "description": BeautifulSoup(job.get("jobDescription", ""), "html.parser").get_text()[:500],
                "source":      "Jobicy",
                "date_found":  datetime.date.today().isoformat(),
                "company_url": job.get("companyUrl", ""),
            })
    except Exception as e:
        print(f"    ⚠️  Jobicy error: {e}")
    return jobs


def scrape_dice(job_title: str, max_results: int = 10) -> list[dict]:
    """Scrape Dice.com — major US job board."""
    print(f"  🌐 Scraping Dice for: {job_title}")
    jobs = []
    try:
        query = job_title.replace(" ", "%20")
        url = f"https://job-search-api.svc.dhigroupinc.com/job-search?query={query}&countryCode2=US&page=1&pageSize={max_results}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        data = resp.json()
        for job in data.get("data", [])[:max_results]:
            jobs.append({
                "title":       job.get("title", ""),
                "company":     job.get("companyName", ""),
                "location":    f"{job.get('jobLocation', {}).get('city', '')}, {job.get('jobLocation', {}).get('state', '')}",
                "url":         f"https://www.dice.com/job-detail/{job.get('id', '')}",
                "description": job.get("summary", "")[:500],
                "source":      "Dice",
                "date_found":  datetime.date.today().isoformat(),
                "company_url": "",
            })
    except Exception as e:
        print(f"    ⚠️  Dice error: {e}")
    return jobs


def scrape_monster(job_title: str, max_results: int = 10) -> list[dict]:
    """Scrape Monster.com — major job board."""
    print(f"  🌐 Scraping Monster for: {job_title}")
    jobs = []
    try:
        query = job_title.replace(" ", "%20")
        url = f"https://www.monster.com/jobs/search?q={query}&page=1"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, "html.parser")
        job_cards = soup.find_all("div", class_="job-card")[:max_results]
        for card in job_cards:
            title_elem = card.find("h3", class_="job-card-title")
            company_elem = card.find("span", class_="job-card-company-name")
            location_elem = card.find("span", class_="job-card-location")
            link_elem = card.find("a", class_="job-card-link")
            desc_elem = card.find("p", class_="job-card-description")

            jobs.append({
                "title":       title_elem.get_text(strip=True) if title_elem else "",
                "company":     company_elem.get_text(strip=True) if company_elem else "",
                "location":    location_elem.get_text(strip=True) if location_elem else "",
                "url":         link_elem.get("href") if link_elem else "",
                "description": desc_elem.get_text(strip=True)[:500] if desc_elem else "",
                "source":      "Monster",
                "date_found":  datetime.date.today().isoformat(),
                "company_url": "",
            })
    except Exception as e:
        print(f"    ⚠️  Monster error: {e}")
    return jobs


# ─────────────────────────────────────────────
#  📧  EMAIL FINDER
# ─────────────────────────────────────────────


def get_domain_from_url(url: str) -> str:
    """Extract domain from a URL."""
    if not url:
        return ""
    url = url.replace("https://", "").replace("http://", "").replace("www.", "")
    return url.split("/")[0].strip()


def guess_emails(company: str, company_url: str) -> list[str]:
    """
    Generate a list of likely email addresses for a company.
    These are guesses — you verify manually before sending.
    """
    domain = get_domain_from_url(company_url)
    if not domain or len(domain) < 4:
        # Try to build domain from company name
        clean = re.sub(r'[^a-zA-Z0-9]', '', company.lower().split()[0])
        domain = f"{clean}.com"

    # Generate HR/careers emails (these are safe generic guesses)
    emails = [
        f"careers@{domain}",
        f"hr@{domain}",
    ]
    return emails


def scrape_emails_from_url(url: str) -> list[str]:
    """
    Scrape emails from a given URL.
    """
    if not url:
        return []
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text()
        # Find all emails using regex
        email_pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
        emails = re.findall(email_pattern, text)
        # Filter valid emails
        valid_emails = [email for email in emails if verify_email_syntax(email)]
        # Remove duplicates and return
        return list(set(valid_emails))
    except Exception as e:
        print(f"    ⚠️  Error scraping {url}: {e}")
        return []


def get_emails_from_hunter(domain: str) -> list[str]:
    """
    Get emails from Hunter.io domain search.
    """
    if not HUNTER_API_KEY or not domain:
        return []
    try:
        url = f"https://api.hunter.io/v2/domain-search?domain={domain}&api_key={HUNTER_API_KEY}"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        emails = []
        for email_data in data.get("data", {}).get("emails", []):
            email = email_data.get("value")
            if email and verify_email_syntax(email):
                emails.append(email)
        return emails[:5]  # limit to 5
    except Exception as e:
        print(f"    ⚠️  Hunter.io error for {domain}: {e}")
        return []


def search_linkedin_professionals(job_title: str, company: str) -> list[str]:
    """
    Search LinkedIn for professionals (HR/hiring managers) via Google search.
    Uses site:linkedin.com search to find relevant profiles and emails.
    """
    emails = []
    try:
        # Search for HR professionals at the company
        search_queries = [
            f"site:linkedin.com {company} HR hiring manager",
            f"site:linkedin.com {company} recruiting",
            f"site:linkedin.com {company} \"recruiter\" email",
        ]
        
        for query in search_queries:
            try:
                # Use Google Custom Search via requests with proper headers
                google_url = f"https://www.google.com/search?q={query.replace(' ', '+')}"
                resp = requests.get(google_url, headers=HEADERS, timeout=10)
                soup = BeautifulSoup(resp.text, "html.parser")
                
                # Extract emails from search results
                email_pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
                found_emails = re.findall(email_pattern, resp.text)
                emails.extend([e for e in found_emails if verify_email_syntax(e)])
                time.sleep(0.5)  # be respectful
            except Exception as e:
                continue
        
        return list(set(emails))[:5]  # deduplicate and limit
    except Exception as e:
        print(f"    ⚠️  LinkedIn search error: {e}")
        return []


def search_google_for_emails(job_title: str, company: str) -> list[str]:
    """
    Search Google for professional emails related to job title and company.
    Targets HR, hiring managers, recruiters, etc.
    """
    emails = []
    try:
        search_queries = [
            f"{company} {job_title} HR email contact",
            f"{company} hiring manager {job_title} email",
            f"{company} \"careers@\" OR \"hr@\" OR \"recruiting@\"",
            f"{company} recruiter contact email",
        ]
        
        for query in search_queries:
            try:
                # Use Google search
                google_url = f"https://www.google.com/search?q={query.replace(' ', '+')}"
                resp = requests.get(google_url, headers=HEADERS, timeout=10)
                
                # Extract emails from search results
                email_pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
                found_emails = re.findall(email_pattern, resp.text)
                emails.extend([e for e in found_emails if verify_email_syntax(e)])
                time.sleep(0.5)  # be respectful
            except Exception as e:
                continue
        
        return list(set(emails))[:8]  # deduplicate and limit to 8
    except Exception as e:
        print(f"    ⚠️  Google search error: {e}")
        return []


def search_professional_networks(company: str, job_title: str) -> list[str]:
    """
    Search multiple professional networks and job boards for emails.
    Looks for contact pages, careers pages, and team listings.
    """
    emails = []
    domain = get_domain_from_url(f"https://{company.lower().replace(' ', '')}.com")
    
    # Common professional contact page paths
    contact_paths = [
        "/careers",
        "/about/team",
        "/contact",
        "/company",
        "/about",
        "/jobs",
        "/hr",
        "/recruiting",
    ]
    
    for path in contact_paths:
        try:
            # Try common company domain structures
            urls_to_try = [
                f"https://{domain}{path}",
                f"https://www.{domain}{path}",
            ]
            
            for url in urls_to_try:
                try:
                    resp = requests.get(url, headers=HEADERS, timeout=5)
                    # Extract emails
                    email_pattern = r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
                    found_emails = re.findall(email_pattern, resp.text)
                    emails.extend([e for e in found_emails if verify_email_syntax(e)])
                except:
                    continue
            time.sleep(0.3)
        except Exception as e:
            continue
    
    return list(set(emails))[:10]


def verify_email_syntax(email: str) -> bool:
    """Basic syntax check."""
    pattern = r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


# ─────────────────────────────────────────────
#  🤖  AI COVER LETTER GENERATOR (Groq — free)
# ─────────────────────────────────────────────

def generate_cover_letter(job: dict) -> str:
    """Use Groq's free API to write a tailored cover letter."""
    if not USE_AI or GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
        return f"Dear Hiring Team at {job['company']},\n\nI am writing to express my strong interest in the {job['title']} position at {job['company']}.\n\n[AI cover letter generation disabled - using default template]\n\nI bring significant expertise in data engineering, DevOps, and cloud technologies. With experience in both startups and multinational companies, I specialize in building systems from scratch and developing end-to-end pipelines.\n\nMy background includes successfully delivering R&D projects and proofs of concept (POCs) in:\n• Databricks & Cloud platforms (AWS, Azure, GCP)\n• ETL pipelines & data architecture\n• Languages: Python, SQL, Data Science and AI tools\n• Infrastructure tools: Airflow, Terraform\n• Startup & MNC experience in R&D projects\n\nI am confident that my skills and experience align closely with what you're looking for. Please find my resume attached for your review.\n\nI am available to start immediately and would welcome the opportunity to discuss how my background can contribute to your team.\n\nBest regards,\n{YOUR_NAME}"

    try:
        from groq import Groq
        client = Groq(api_key=GROQ_API_KEY)

        prompt = f"""Write a SHORT, professional cover letter email body (3 paragraphs max) for this job.

JOB TITLE: {job['title']}
COMPANY: {job['company']}
JOB DESCRIPTION (excerpt): {job['description'][:400]}

MY RESUME:
{RESUME_TEXT}

Rules:
- Start with "Dear Hiring Team at {job['company']},"
- Be specific to this role — mention 1-2 skills that match the job description
- Keep it under 200 words — recruiters hate long emails
- End with "Best regards,\n{YOUR_NAME}\n{YOUR_EMAIL}\n{YOUR_PHONE}"
- Do NOT use generic phrases like "I am a passionate professional"
"""
        response = client.chat.completions.create(
            model="mixtral-8x7b-32768",   # Updated: Mixtral model for cover letters
            messages=[{"role": "user", "content": prompt}],
            max_tokens=400,
            temperature=0.7,
        )
        return response.choices[0].message.content.strip()

    except Exception as e:
        print(f"    ⚠️  AI error: {e}")
        return f"Dear Hiring Team at {job['company']},\n\nI am writing to express my interest in the {job['title']} position.\n\nBest regards,\n{YOUR_NAME}"


# ─────────────────────────────────────────────
#  📊  EXCEL OUTPUT
# ─────────────────────────────────────────────

def write_to_excel(jobs_data: list[dict], filename: str):
    """Write all job data to a nicely formatted Excel file."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Applications tracker ──
    ws1 = wb.active
    ws1.title = "Applications"

    headers = [
        "Status", "Company", "Job Title", "Location", "Source",
        "Apply URL", "Email 1", "Email 2", "Email 3", "Email 4", "Email 5",
        "Cover Letter", "Date Found", "Notes"
    ]

    # Style header row
    header_fill = PatternFill("solid", fgColor="2D3748")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Status options with colors
    status_colors = {
        "🔵 To Apply":    "EBF8FF",
        "✅ Applied":     "F0FFF4",
        "⏳ Interviewing":"FFFFF0",
        "❌ Rejected":    "FFF5F5",
    }

    # Write each job
    for row_idx, job in enumerate(jobs_data, 2):
        row_fill = PatternFill("solid", fgColor="F7FAFC" if row_idx % 2 == 0 else "FFFFFF")

        ws1.cell(row=row_idx, column=1,  value="🔵 To Apply").fill = PatternFill("solid", fgColor="EBF8FF")
        ws1.cell(row=row_idx, column=2,  value=job.get("company", ""))
        ws1.cell(row=row_idx, column=3,  value=job.get("title", ""))
        ws1.cell(row=row_idx, column=4,  value=job.get("location", ""))
        ws1.cell(row=row_idx, column=5,  value=job.get("source", ""))
        ws1.cell(row=row_idx, column=6,  value=job.get("url", ""))
        emails = job.get("guessed_emails", [])
        ws1.cell(row=row_idx, column=7,  value=emails[0] if len(emails) > 0 else "")
        ws1.cell(row=row_idx, column=8,  value=emails[1] if len(emails) > 1 else "")
        ws1.cell(row=row_idx, column=9,  value=emails[2] if len(emails) > 2 else "")
        ws1.cell(row=row_idx, column=10, value=emails[3] if len(emails) > 3 else "")
        ws1.cell(row=row_idx, column=11, value=emails[4] if len(emails) > 4 else "")
        ws1.cell(row=row_idx, column=12, value=job.get("cover_letter", "")).alignment = Alignment(wrap_text=True)
        ws1.cell(row=row_idx, column=13, value=job.get("date_found", ""))
        ws1.cell(row=row_idx, column=14, value="")  # Notes column

        # Alternate row shading
        for col in [2, 3, 4, 5, 11, 12]:
            ws1.cell(row=row_idx, column=col).fill = row_fill

    # Column widths
    col_widths = [14, 22, 28, 16, 12, 40, 30, 30, 30, 30, 30, 60, 14, 20]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: Email Templates ──
    ws2 = wb.create_sheet("Email Templates")
    ws2["A1"] = "📧 Email Subject Line Templates"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A3"] = "Subject line ideas:"
    templates = [
        "Application: {Job Title} — {Your Name}",
        "Interested in {Job Title} role at {Company}",
        "{Your Name} | {Job Title} Application",
        "Re: {Job Title} Opening at {Company}",
    ]
    for i, t in enumerate(templates, 4):
        ws2.cell(row=i, column=1, value=t)

    ws2["A10"] = "📌 How to use this sheet:"
    ws2["A10"].font = Font(bold=True)
    instructions = [
        "1. Filter the Applications tab by 'Status = To Apply'",
        "2. Open the job URL to confirm it's still active",
        "3. Copy the Cover Letter column into your email body",
        "4. Try Email 1 first — careers@ or hr@ usually works, or scraped emails",
        "5. Update Status to '✅ Applied' after sending",
        "6. Add notes like 'Sent via LinkedIn' in the Notes column",
    ]
    for i, inst in enumerate(instructions, 11):
        ws2.cell(row=i, column=1, value=inst)

    ws2.column_dimensions["A"].width = 60

    wb.save(filename)
    print(f"\n  ✅ Saved to: {filename}")


# ─────────────────────────────────────────────
#  🚀  MAIN RUNNER
# ─────────────────────────────────────────────

def main():
    print("\n" + "="*55)
    print("  🤖 JOB HUNTER AGENT STARTING")
    print(f"  Looking for: {', '.join(JOB_TITLES)}")
    print(f"  Location: {LOCATION}")
    print("="*55 + "\n")

    all_jobs = []
    seen_urls = set()

    # 1. Scrape jobs from all sources
    per_source = max(10, MAX_JOBS // len(JOB_TITLES))
    for title in JOB_TITLES:
        for scraper in [scrape_remotive, scrape_arbeitnow, scrape_jobicy, scrape_dice, scrape_monster]:
            jobs = scraper(title, max_results=per_source)
            for job in jobs:
                if job["url"] not in seen_urls and job["title"]:
                    seen_urls.add(job["url"])
                    all_jobs.append(job)
            time.sleep(1)  # be polite

    print(f"\n  📋 Found {len(all_jobs)} unique jobs\n")

    # 2. Find emails + generate cover letters
    for i, job in enumerate(all_jobs):
        print(f"  [{i+1}/{len(all_jobs)}] {job['company']} — {job['title']}")

        # Scrape emails from company website and job posting
        scraped = []
        urls_to_scrape = [job.get("company_url", ""), job.get("url", "")]
        for url in urls_to_scrape:
            if url:
                scraped.extend(scrape_emails_from_url(url))
                time.sleep(0.5)  # be polite

        # Get emails from Hunter.io
        domain = get_domain_from_url(job.get("company_url", ""))
        hunter_emails = get_emails_from_hunter(domain)

        # Search LinkedIn for HR professionals
        print(f"    🔍 Searching LinkedIn for HR professionals...")
        linkedin_emails = search_linkedin_professionals(job["title"], job["company"])

        # Search Google for emails
        print(f"    🔍 Searching Google for contact emails...")
        google_emails = search_google_for_emails(job["title"], job["company"])

        # Search professional networks (careers pages, team pages, etc)
        print(f"    🔍 Searching professional networks...")
        network_emails = search_professional_networks(job["company"], job["title"])

        # Guess emails (last resort)
        guessed = guess_emails(job["company"], job.get("company_url", ""))

        # Combine and deduplicate
        all_emails = list(set(scraped + hunter_emails + linkedin_emails + google_emails + network_emails + guessed))
        job["guessed_emails"] = all_emails[:5]  # limit to 5 emails

        # AI cover letter
        if USE_AI:
            print(f"    🤖 Writing cover letter...")
            job["cover_letter"] = generate_cover_letter(job)
            time.sleep(0.5)  # rate limit
        else:
            job["cover_letter"] = f"Dear Hiring Team at {job['company']},\n\nI am writing to apply for the {job['title']} position.\n\nBest regards,\n{YOUR_NAME}"

    # 3. Write to Excel
    print(f"\n  📊 Writing to Excel...")
    write_to_excel(all_jobs, OUTPUT_FILE)

    # 4. Summary
    print("\n" + "="*55)
    print(f"  ✅ DONE! {len(all_jobs)} jobs saved to {OUTPUT_FILE}")
    print("\n  NEXT STEPS:")
    print("  1. Open job_applications.xlsx")
    print("  2. Review jobs — delete irrelevant ones")
    print("  3. Edit cover letters as needed")
    print("  4. Send emails manually (copy/paste from Excel)")
    print("  5. Update Status column as you apply")
    print("="*55 + "\n")


if __name__ == "__main__":
    main()